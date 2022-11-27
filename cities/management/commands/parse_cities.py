import os
from pathlib import Path
from django.db import models
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from cities.models import City

class Command(BaseCommand):
    help = 'Load data from csv'

    def handle(self, *args, **options):
        City.objects.all().delete()
        print('table dropped')

        base_dir = Path(__file__).resolve().parent.parent.parent.parent
        book_path = os.path.join(base_dir, 'cities/city_data/urbanspatial-hist-urban-pop-3700bc-ad2000-xlsx.xlsx')
        book = load_workbook(book_path)
        sheet = book['Historical Urban Population']
        print(sheet.title)
        max_row_num = sheet.max_row
        max_col_num = sheet.max_column
        print(max_row_num)
        print(max_col_num)

        # placeholder variables for city object
        city = "temp_name"
        otherName = "None"
        country = "temp_country"
        latitude = 0.0
        longitude = 0.0
        year = 1111
        pop = 111
        city_id = "temp_id"

        # as this is a spreadsheet and not a csv file, we need to iterate cell by cell over a range of cells
        # skip first row as headers, and skip first column as we don't need it
        for i in range(2, max_row_num+1):

            for j in range(2, max_col_num+1):
                cell_obj=sheet.cell(row=i, column=j)
                if cell_obj.column_letter=='B':
                    city = cell_obj.value
                if cell_obj.column_letter=='C':
                    if cell_obj.value is not None:
                        otherName = cell_obj.value
                if cell_obj.column_letter=='D':
                    country = cell_obj.value
                if cell_obj.column_letter=='E':
                    latitude = cell_obj.value
                if cell_obj.column_letter=='F':
                    longitude = cell_obj.value
                if cell_obj.column_letter=='H':
                    year = cell_obj.value
                if cell_obj.column_letter=='I':
                    pop = cell_obj.value
                if cell_obj.column_letter=='J':
                    city_id = cell_obj.value
                

                print(cell_obj.value, end='|')
            # end loop so construct city object
            city = City.objects.create(city=city,otherName=otherName,country=country,latitude=latitude,longitude=longitude,year=year,pop=pop,city_id=city_id)
            city.save()
            print(' saved ')
            print('\n')


            
        